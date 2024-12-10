import sqlite3
from openpyxl import load_workbook

# 1. Excel faylini yuklash
file_path = 'users_data.xlsx'  # Excel fayl nomi
sheet_name = 'Telegram Bot Leedlari'    # Excel varaq nomi

# Excel faylini yuklash
wb = load_workbook(file_path)
ws = wb[sheet_name]

# 2. Ma'lumotlar bazasiga ulanish
conn = sqlite3.connect('bot_data.db')  # SQLite bazangiz
cursor = conn.cursor()

# Jadval yaratish (agar mavjud bo'lmasa)
cursor.execute("""
    CREATE TABLE IF NOT EXISTS user_full_data (
        ism TEXT NULL,
        tg_id TEXT UNIQUE NULL,
        phone_number TEXT NULL,
        joined_data TEXT NULL
    )
""")

# 3. Excel fayldan ma'lumotlarni o'qish va Database-ga qo'shish
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]  # Sarlavha qatori (FISH, Telegram ID, Telefon Raqami, Ro'yxatdan o'tgan sanasi)

for row in rows[1:]:  # Birinchi sarlavha qatordan tashqari qatorlarni o'qish
    try:
        cursor.execute("""
            INSERT INTO user_full_data (ism, tg_id, phone_number, joined_data)
            VALUES (?, ?, ?, ?)
        """, (row[0], row[1], row[2], row[3]))
    except sqlite3.IntegrityError:
        print(f"Ma'lumot allaqachon mavjud: {row[1]}")  # tg_id takrorlanganda xabar chiqarish

# O'zgarishlarni saqlash va ulanishni yopish
conn.commit()
conn.close()

print("Ma'lumotlar muvaffaqiyatli qo'shildi!")
