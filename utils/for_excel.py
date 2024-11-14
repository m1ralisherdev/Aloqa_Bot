from openpyxl import Workbook
from database import get_all_data
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

async def create_excel():
    data = get_all_data()
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Telegram Bot Leedlari'
    sheet.column_dimensions['A'].width = 35  
    sheet.column_dimensions['B'].width = 35  
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 35  

    headers = ['FISH', 'Telegram ID', 'Telefon Raqami', "Ro'yxatdan o'tgan sanasi"]
    sheet.append(headers)  # Sarlavhalarni qo'shish

    # Ma'lumotlarni yozish
    for row_num, row in enumerate(data, 2):
        for col_num, cell_value in enumerate(row, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Sarlavha kataklarini formatlash
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(name='Arial', size=12, bold=True, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ustun kengligini hisobga olish

    workbook.save('users_data.xlsx')

